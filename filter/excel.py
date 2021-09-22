import json
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


f = open('final_vals_all.json', 'r', encoding='utf-8')
final_dic = json.load(f)
f.close()

wb = Workbook()

dest_filename = 'vals_all.xlsx'

ws1 = wb.active

ws1.title = "sheet_1"

row = 2
for a in final_dic.keys():
    _ = ws1.cell(column=1, row=row, value=a)
    _ = ws1.cell(column=2, row=row, value=final_dic[a])
#    _ = ws1.cell(column=3, row=row, value=final_dic[a]['weighted'])
    row = row + 1

wb.save(filename= dest_filename)

